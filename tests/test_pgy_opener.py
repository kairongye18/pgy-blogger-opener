import sys
import sqlite3
import unittest
from ssl import SSLCertVerificationError
from pathlib import Path
from unittest.mock import Mock, patch
from urllib.error import URLError

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from pgy_opener import (  # noqa: E402
    BrowserOpener,
    Creator,
    _chromium_blank_window_script,
    chrome_profile_new_window_command,
    build_pgy_detail_url,
    default_browser_bundle_id_from_handlers,
    extract_profile_user_id,
    find_chrome_profile_with_xhs_cookies,
    load_creators,
    macos_browser_application_name,
    open_creator_tabs,
    resolve_short_link,
    select_batch,
    windows_chrome_executable,
)


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["提交时间", "小红书昵称（必填）", "粉丝数量", "主页链接（必填）"])
    ws.append(["2026-06-09", "肥肥泡芙", "1.1w", "https://www.xiaohongshu.com/user/profile/6425da410000000011022fd3?x=1"])
    ws.append(["2026-06-09", "金条爱吃自助", "3200", "https://xhslink.com/m/86GaCLuKXOB"])
    ws.append(["2026-06-09", "", "999", "https://www.xiaohongshu.com/user/profile/skipme"])
    wb.save(path)


class PgyOpenerTests(unittest.TestCase):
    def test_extract_profile_user_id_from_xiaohongshu_profile_url(self):
        self.assertEqual(
            extract_profile_user_id(
                "https://www.xiaohongshu.com/user/profile/6425da410000000011022fd3?xsec_token=abc"
            ),
            "6425da410000000011022fd3",
        )

    def test_extract_profile_user_id_from_login_redirect_url(self):
        self.assertEqual(
            extract_profile_user_id(
                "https://www.xiaohongshu.com/login?redirectPath="
                "https%3A%2F%2Fwww.xiaohongshu.com%2Fuser%2Fprofile%2F6425da410000000011022fd3%3Fx%3D1"
            ),
            "6425da410000000011022fd3",
        )

    def test_build_pgy_detail_url(self):
        self.assertEqual(
            build_pgy_detail_url("6425da410000000011022fd3"),
            "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/"
            "6425da410000000011022fd3?fromRoute=Advertiser_Kol",
        )

    def test_load_creators_detects_headers_and_skips_blank_names(self):
        with self.subTest("workbook"):
            import tempfile

            with tempfile.TemporaryDirectory() as tmp:
                workbook = Path(tmp) / "creators.xlsx"
                make_workbook(workbook)

                creators = load_creators(workbook, resolve_short_links=False)

        self.assertEqual([creator.name for creator in creators], ["肥肥泡芙", "金条爱吃自助"])
        self.assertEqual(creators[0].profile_user_id, "6425da410000000011022fd3")
        self.assertIsNone(creators[1].profile_user_id)
        self.assertEqual(creators[1].error, "短链未解析，缺少小红书 userId")

    def test_load_creators_resolves_short_links_with_injected_resolver(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "creators.xlsx"
            make_workbook(workbook)

            creators = load_creators(
                workbook,
                resolve_short_links=True,
                resolver=lambda url: "https://www.xiaohongshu.com/user/profile/5cd7eb39000000001203a67d?abc=1",
            )

        self.assertEqual(creators[1].profile_user_id, "5cd7eb39000000001203a67d")
        self.assertTrue(creators[1].pgy_url.endswith("/5cd7eb39000000001203a67d?fromRoute=Advertiser_Kol"))
        self.assertIsNone(creators[1].error)

    def test_load_creators_only_resolves_requested_batch(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "creators.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["小红书昵称（必填）", "主页链接（必填）"])
            ws.append(["第一个", "https://xhslink.com/m/first"])
            ws.append(["第二个", "https://xhslink.com/m/second"])
            ws.append(["第三个", "https://xhslink.com/m/third"])
            wb.save(workbook)

            resolved_urls = []

            def resolver(url: str) -> str:
                resolved_urls.append(url)
                return f"https://www.xiaohongshu.com/user/profile/{url.rsplit('/', 1)[-1]}"

            creators = load_creators(workbook, start=2, count=1, resolver=resolver)

        self.assertEqual([creator.name for creator in creators], ["第二个"])
        self.assertEqual(resolved_urls, ["https://xhslink.com/m/second"])

    def test_resolve_short_link_retries_without_certificate_verification_for_local_python_cert_errors(self):
        failed_call = URLError(SSLCertVerificationError("CERTIFICATE_VERIFY_FAILED"))
        successful_response = Mock()
        successful_response.__enter__ = Mock(return_value=successful_response)
        successful_response.__exit__ = Mock(return_value=None)
        successful_response.geturl.return_value = "https://www.xiaohongshu.com/user/profile/5cd7eb39000000001203a67d"

        with patch("pgy_opener.urlopen", side_effect=[failed_call, successful_response]) as mocked_urlopen:
            resolved = resolve_short_link("https://xhslink.com/m/86GaCLuKXOB")

        self.assertEqual(resolved, "https://www.xiaohongshu.com/user/profile/5cd7eb39000000001203a67d")
        self.assertEqual(mocked_urlopen.call_count, 2)
        self.assertIn("context", mocked_urlopen.call_args.kwargs)

    def test_open_creator_tabs_starts_a_new_window_for_each_group(self):
        creators = [
            Creator(row_number=i + 2, name=f"博主{i + 1}", profile_link="", profile_user_id=f"user{i + 1}")
            for i in range(12)
        ]
        calls = []
        opener = BrowserOpener(
            open_blank_window=lambda: calls.append(("blank", None)),
            open_url_in_current_tab=lambda url: calls.append(("current", url)),
            open_url_in_new_tab=lambda url: calls.append(("tab", url)),
            open_url_in_new_window=lambda url: calls.append(("window", url)),
            name="测试浏览器",
        )

        with patch("pgy_opener.time.sleep"):
            open_creator_tabs(creators, delay=0, window_size=10, progress=lambda message: None, browser_opener=opener)

        self.assertEqual(calls[0][0], "window")
        self.assertIn("/user1?", calls[0][1])
        self.assertEqual(calls[9][0], "tab")
        self.assertIn("/user10?", calls[9][1])
        self.assertEqual(calls[10][0], "window")
        self.assertIn("/user11?", calls[10][1])
        self.assertEqual(calls[-1][0], "tab")
        self.assertIn("/user12?", calls[-1][1])

    def test_open_creator_tabs_sends_messages_to_progress_callback(self):
        creators = [Creator(row_number=2, name="肥肥泡芙", profile_link="", profile_user_id="user1")]
        messages = []
        opener = BrowserOpener(
            open_blank_window=lambda: None,
            open_url_in_current_tab=lambda url: None,
            open_url_in_new_tab=lambda url: None,
            open_url_in_new_window=lambda url: None,
            name="测试浏览器",
        )

        open_creator_tabs(creators, delay=0, progress=messages.append, browser_opener=opener)

        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0], "使用浏览器：测试浏览器")
        self.assertIn("[打开新窗口] 第 2 行 肥肥泡芙", messages[1])

    def test_default_browser_bundle_id_prefers_https_handler(self):
        handlers = [
            {"LSHandlerContentType": "com.apple.default-app.web-browser", "LSHandlerRoleAll": "com.apple.safari"},
            {"LSHandlerURLScheme": "https", "LSHandlerRoleAll": "com.brave.Browser"},
        ]

        self.assertEqual(default_browser_bundle_id_from_handlers(handlers), "com.brave.browser")

    def test_default_browser_bundle_id_falls_back_to_default_browser_content_type(self):
        handlers = [
            {"LSHandlerContentType": "com.apple.default-app.web-browser", "LSHandlerRoleViewer": "com.microsoft.edgemac"}
        ]

        self.assertEqual(default_browser_bundle_id_from_handlers(handlers), "com.microsoft.edgemac")

    def test_macos_browser_application_name_supports_common_logged_in_browsers(self):
        self.assertEqual(macos_browser_application_name("com.google.chrome"), "Google Chrome")
        self.assertEqual(macos_browser_application_name("com.microsoft.edgemac"), "Microsoft Edge")
        self.assertEqual(macos_browser_application_name("com.brave.browser"), "Brave Browser")
        self.assertEqual(macos_browser_application_name("company.thebrowser.browser"), "Arc")
        self.assertEqual(macos_browser_application_name("com.apple.safari"), "Safari")
        self.assertIsNone(macos_browser_application_name("unknown.browser"))

    def test_chromium_blank_window_uses_keyboard_shortcut_to_keep_active_profile(self):
        script = _chromium_blank_window_script("Google Chrome")

        self.assertIn('tell application "Google Chrome"', script)
        self.assertIn("pgy.xiaohongshu.com", script)
        self.assertIn("set matchedWindow to browserWindow", script)
        self.assertIn("set index of matchedWindow to 1", script)
        self.assertIn('tell application "System Events"', script)
        self.assertIn('keystroke "n" using command down', script)
        self.assertNotIn("make new window", script)

    def test_find_chrome_profile_with_xhs_cookies_returns_profile_directory_name(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            chrome_dir = Path(tmp)
            cookies = chrome_dir / "Profile 2" / "Network" / "Cookies"
            cookies.parent.mkdir(parents=True)
            connection = sqlite3.connect(cookies)
            connection.execute("create table cookies(host_key text)")
            connection.execute("insert into cookies(host_key) values (?)", (".xiaohongshu.com",))
            connection.commit()
            connection.close()

            self.assertEqual(find_chrome_profile_with_xhs_cookies(chrome_dir), "Profile 2")

    def test_find_chrome_profile_with_xhs_cookies_ignores_profiles_without_xhs_cookie(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            chrome_dir = Path(tmp)
            cookies = chrome_dir / "Default" / "Network" / "Cookies"
            cookies.parent.mkdir(parents=True)
            connection = sqlite3.connect(cookies)
            connection.execute("create table cookies(host_key text)")
            connection.execute("insert into cookies(host_key) values (?)", (".example.com",))
            connection.commit()
            connection.close()

            self.assertIsNone(find_chrome_profile_with_xhs_cookies(chrome_dir))

    def test_chrome_profile_new_window_command_is_atomic_and_profile_specific(self):
        command = chrome_profile_new_window_command(
            "Profile 1",
            "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/user1",
        )

        self.assertEqual(
            command[0],
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        )
        self.assertIn("--profile-directory=Profile 1", command)
        self.assertIn("--new-window", command)
        self.assertEqual(command[-1], "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/user1")

    def test_windows_chrome_executable_finds_per_user_install(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            local_app_data = Path(tmp)
            chrome = local_app_data / "Google" / "Chrome" / "Application" / "chrome.exe"
            chrome.parent.mkdir(parents=True)
            chrome.write_bytes(b"MZ")

            executable = windows_chrome_executable({"LOCALAPPDATA": str(local_app_data)})

        self.assertEqual(executable, chrome)

    def test_windows_browser_opener_uses_logged_in_chrome_profile_and_new_window_flag(self):
        chrome = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")
        detail_url = "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/user1"

        with (
            patch("pgy_opener.sys.platform", "win32"),
            patch("pgy_opener.windows_chrome_executable", return_value=chrome),
            patch("pgy_opener.find_chrome_profile_with_xhs_cookies", return_value="Profile 2"),
            patch("pgy_opener.subprocess.Popen") as popen,
        ):
            opener = __import__("pgy_opener").create_browser_opener()
            opener.open_url_in_new_window(detail_url)
            opener.open_url_in_new_tab(detail_url)

        new_window_command = popen.call_args_list[0].args[0]
        new_tab_command = popen.call_args_list[1].args[0]
        self.assertEqual(new_window_command[0], str(chrome))
        self.assertIn("--profile-directory=Profile 2", new_window_command)
        self.assertIn("--new-window", new_window_command)
        self.assertEqual(new_window_command[-1], detail_url)
        self.assertEqual(new_tab_command[0], str(chrome))
        self.assertIn("--profile-directory=Profile 2", new_tab_command)
        self.assertNotIn("--new-window", new_tab_command)
        self.assertEqual(opener.name, "Google Chrome（Profile 2）")

    def test_select_batch_is_one_based_and_limited(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "creators.xlsx"
            make_workbook(workbook)
            creators = load_creators(workbook, resolve_short_links=False)

        self.assertEqual([creator.name for creator in select_batch(creators, start=2, count=10)], ["金条爱吃自助"])

    def test_load_creators_requires_nickname_and_profile_link_columns(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "bad.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["昵称", "报价"])
            ws.append(["肥肥泡芙", "1000"])
            wb.save(workbook)

            with self.assertRaisesRegex(ValueError, "主页链接"):
                load_creators(workbook)


if __name__ == "__main__":
    unittest.main()
