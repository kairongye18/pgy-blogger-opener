#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import plistlib
import shutil
import sqlite3
import subprocess
import ssl
import sys
import time
import tempfile
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import parse_qs, urlparse
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

PGY_DETAIL_PREFIX = "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail"
TOOL_VERSION = "2026-06-15.14"


@dataclass
class BrowserOpener:
    open_blank_window: Callable[[], None]
    open_url_in_current_tab: Callable[[str], None]
    open_url_in_new_tab: Callable[[str], None]
    open_url_in_new_window: Callable[[str], None]
    name: str = "默认浏览器"


@dataclass
class Creator:
    row_number: int
    name: str
    profile_link: str
    resolved_link: str | None = None
    profile_user_id: str | None = None
    error: str | None = None

    @property
    def pgy_url(self) -> str | None:
        if not self.profile_user_id:
            return None
        return build_pgy_detail_url(self.profile_user_id)


def build_pgy_detail_url(user_id: str) -> str:
    return f"{PGY_DETAIL_PREFIX}/{user_id}?fromRoute=Advertiser_Kol"


def extract_profile_user_id(url: str) -> str | None:
    parsed = urlparse(url)
    marker = "/user/profile/"
    if marker in parsed.path:
        return parsed.path.split(marker, 1)[1].split("/", 1)[0] or None

    query = parse_qs(parsed.query)
    redirect_path = query.get("redirectPath", [""])[0]
    if redirect_path:
        redirected_user_id = extract_profile_user_id(redirect_path)
        if redirected_user_id:
            return redirected_user_id

    for key in ("appuid", "user_id", "userId"):
        value = query.get(key, [""])[0]
        if value:
            return value
    return None


def resolve_short_link(url: str, timeout: int = 15) -> str:
    request = Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 Chrome/125 Safari/537.36"
            )
        },
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            return response.geturl()
    except URLError as exc:
        if not _is_certificate_error(exc):
            raise

    unverified_context = ssl._create_unverified_context()
    with urlopen(request, timeout=timeout, context=unverified_context) as response:
        return response.geturl()


def _is_certificate_error(exc: URLError) -> bool:
    reason = getattr(exc, "reason", exc)
    return isinstance(reason, ssl.SSLCertVerificationError) or "CERTIFICATE_VERIFY_FAILED" in str(exc)


def find_column(headers: list[str], candidates: Iterable[str]) -> int:
    for index, header in enumerate(headers):
        normalized = header.replace(" ", "")
        if any(candidate in normalized for candidate in candidates):
            return index
    raise ValueError(f"未找到必要列：{', '.join(candidates)}")


def load_creators(
    workbook_path: str | Path,
    *,
    sheet_name: str | None = None,
    resolve_short_links: bool = True,
    resolver: Callable[[str], str] = resolve_short_link,
    start: int = 1,
    count: int | None = None,
    progress: Callable[[str], None] | None = None,
) -> list[Creator]:
    if start < 1:
        raise ValueError("--start 必须大于等于 1")
    if count is not None and count < 1:
        raise ValueError("--count 必须大于等于 1")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = ["" if value is None else str(value).strip() for value in next(rows)]
    except StopIteration as exc:
        workbook.close()
        raise ValueError("Excel 文件为空") from exc

    try:
        name_index = find_column(headers, ("小红书昵称", "昵称"))
        link_index = find_column(headers, ("主页链接", "主页"))
    except Exception:
        workbook.close()
        raise

    creators: list[Creator] = []
    valid_creator_number = 0
    for row_number, row in enumerate(rows, start=2):
        name = _cell_text(row, name_index)
        profile_link = _cell_text(row, link_index)
        if not name:
            continue

        valid_creator_number += 1
        if valid_creator_number < start:
            continue
        if count is not None and len(creators) >= count:
            break

        creator = Creator(row_number=row_number, name=name, profile_link=profile_link)
        if not profile_link:
            creator.error = "缺少主页链接"
            creators.append(creator)
            continue

        resolved_link = profile_link
        if "xhslink.com" in profile_link:
            if resolve_short_links:
                try:
                    if progress:
                        progress(f"正在解析第 {row_number} 行 {name} 的短链...")
                    resolved_link = resolver(profile_link)
                except Exception as exc:  # noqa: BLE001 - surface resolver details to the operator.
                    creator.error = f"短链解析失败：{exc}"
            else:
                creator.error = "短链未解析，缺少小红书 userId"

        creator.resolved_link = resolved_link
        creator.profile_user_id = extract_profile_user_id(resolved_link)
        if creator.profile_user_id:
            creator.error = None
        elif not creator.error:
            creator.error = f"主页链接中未找到小红书 userId；最终链接：{resolved_link}"
        creators.append(creator)

    workbook.close()
    return creators


def select_batch(creators: list[Creator], *, start: int, count: int) -> list[Creator]:
    if start < 1:
        raise ValueError("--start 必须大于等于 1")
    if count < 1:
        raise ValueError("--count 必须大于等于 1")
    return creators[start - 1 : start - 1 + count]


def open_creator_tabs(
    creators: list[Creator],
    *,
    delay: float = 0.3,
    dry_run: bool = False,
    window_size: int = 10,
    progress: Callable[[str], None] | None = None,
    browser_opener: BrowserOpener | None = None,
) -> None:
    if window_size < 1:
        raise ValueError("--window-size 必须大于等于 1")

    opener = browser_opener or create_browser_opener()
    if not dry_run:
        _progress(progress, f"使用浏览器：{opener.name}")
    opened_count = 0
    for creator in creators:
        if not creator.pgy_url:
            _progress(progress, f"[跳过] 第 {creator.row_number} 行 {creator.name}: {creator.error}")
            continue

        is_new_window = opened_count % window_size == 0
        action = "新窗口" if is_new_window else "标签页"
        _progress(progress, f"[打开{action}] 第 {creator.row_number} 行 {creator.name}: {creator.pgy_url}")
        if not dry_run:
            if is_new_window:
                opener.open_url_in_new_window(creator.pgy_url)
            else:
                opener.open_url_in_new_tab(creator.pgy_url)
            time.sleep(delay)
        opened_count += 1


def create_browser_opener() -> BrowserOpener:
    if sys.platform == "darwin":
        browser_name = macos_browser_application_name(default_macos_browser_bundle_id())
        if browser_name:
            return _create_macos_browser_opener(browser_name)
        return BrowserOpener(
            open_blank_window=_open_blank_window_macos,
            open_url_in_current_tab=_open_url_in_current_tab_macos,
            open_url_in_new_tab=_open_url_in_new_tab_macos,
            open_url_in_new_window=lambda url: webbrowser.open(url, new=1),
            name="默认浏览器",
        )
    if sys.platform == "win32":
        chrome_executable = windows_chrome_executable()
        if chrome_executable:
            chrome_profile = find_chrome_profile_with_xhs_cookies(_windows_chrome_user_data_dir())
            return _create_windows_chrome_opener(chrome_executable, chrome_profile)
    return BrowserOpener(
        open_blank_window=lambda: webbrowser.open("about:blank", new=1),
        open_url_in_current_tab=lambda url: webbrowser.open(url, new=0),
        open_url_in_new_tab=webbrowser.open_new_tab,
        open_url_in_new_window=lambda url: webbrowser.open(url, new=1),
        name="默认浏览器",
    )


def _create_macos_browser_opener(browser_name: str) -> BrowserOpener:
    if browser_name == "Safari":
        return BrowserOpener(
            open_blank_window=_open_blank_window_safari_macos,
            open_url_in_current_tab=_open_url_in_current_tab_safari_macos,
            open_url_in_new_tab=_open_url_in_new_tab_safari_macos,
            open_url_in_new_window=lambda url: _open_url_in_new_window_safari_macos(url),
            name=browser_name,
        )

    chrome_profile = find_chrome_profile_with_xhs_cookies() if browser_name == "Google Chrome" else None
    blank_window = (
        (lambda: _open_blank_window_chrome_profile_macos(chrome_profile))
        if chrome_profile
        else (lambda: _open_blank_window_chromium_macos(browser_name))
    )
    return BrowserOpener(
        open_blank_window=blank_window,
        open_url_in_current_tab=lambda url: _open_url_in_current_tab_chromium_macos(browser_name, url),
        open_url_in_new_tab=lambda url: _open_url_in_new_tab_chromium_macos(browser_name, url),
        open_url_in_new_window=(
            (lambda url: _open_url_in_new_window_chrome_profile_macos(chrome_profile, url))
            if chrome_profile
            else (lambda url: _open_url_in_new_window_chromium_macos(browser_name, url))
        ),
        name=f"{browser_name}（{chrome_profile}）" if chrome_profile else browser_name,
    )


def windows_chrome_executable(environ: dict[str, str] | None = None) -> Path | None:
    environment = environ if environ is not None else os.environ
    candidates = [
        Path(environment.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(environment.get("PROGRAMFILES", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(environment.get("PROGRAMFILES(X86)", "")) / "Google" / "Chrome" / "Application" / "chrome.exe",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def _windows_chrome_user_data_dir(environ: dict[str, str] | None = None) -> Path:
    environment = environ if environ is not None else os.environ
    return Path(environment.get("LOCALAPPDATA", "")) / "Google" / "Chrome" / "User Data"


def _create_windows_chrome_opener(chrome_executable: Path, profile_name: str | None) -> BrowserOpener:
    def open_chrome(url: str, *, new_window: bool) -> None:
        command = [str(chrome_executable)]
        if profile_name:
            command.append(f"--profile-directory={profile_name}")
        if new_window:
            command.append("--new-window")
        command.append(url)
        subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    return BrowserOpener(
        open_blank_window=lambda: open_chrome("about:blank", new_window=True),
        open_url_in_current_tab=lambda url: open_chrome(url, new_window=False),
        open_url_in_new_tab=lambda url: open_chrome(url, new_window=False),
        open_url_in_new_window=lambda url: open_chrome(url, new_window=True),
        name=f"Google Chrome（{profile_name}）" if profile_name else "Google Chrome",
    )


def default_macos_browser_bundle_id() -> str | None:
    launch_services = Path.home() / "Library" / "Preferences" / "com.apple.LaunchServices" / "com.apple.launchservices.secure.plist"
    try:
        with launch_services.open("rb") as file:
            data = plistlib.load(file)
    except Exception:
        return None
    handlers = data.get("LSHandlers", [])
    if not isinstance(handlers, list):
        return None
    return default_browser_bundle_id_from_handlers(handlers)


def default_browser_bundle_id_from_handlers(handlers: list[dict[str, object]]) -> str | None:
    https_handler = _find_browser_handler(handlers, "LSHandlerURLScheme", "https")
    if https_handler:
        return https_handler
    return _find_browser_handler(handlers, "LSHandlerContentType", "com.apple.default-app.web-browser")


def _find_browser_handler(handlers: list[dict[str, object]], key: str, value: str) -> str | None:
    for handler in reversed(handlers):
        if str(handler.get(key, "")).lower() != value:
            continue
        bundle_id = handler.get("LSHandlerRoleAll") or handler.get("LSHandlerRoleViewer")
        if bundle_id and bundle_id != "__dummy__":
            return str(bundle_id).lower()
    return None


def macos_browser_application_name(bundle_id: str | None) -> str | None:
    if not bundle_id:
        return None
    return {
        "com.google.chrome": "Google Chrome",
        "com.google.chrome.canary": "Google Chrome Canary",
        "com.microsoft.edgemac": "Microsoft Edge",
        "com.microsoft.edgemac.canary": "Microsoft Edge Canary",
        "com.brave.browser": "Brave Browser",
        "company.thebrowser.browser": "Arc",
        "com.apple.safari": "Safari",
        "com.apple.safari.technologypreview": "Safari Technology Preview",
        "com.vivaldi.vivaldi": "Vivaldi",
        "com.operasoftware.opera": "Opera",
    }.get(bundle_id.lower())


def find_chrome_profile_with_xhs_cookies(chrome_user_data_dir: str | Path | None = None) -> str | None:
    chrome_dir = Path(chrome_user_data_dir) if chrome_user_data_dir else _chrome_user_data_dir()
    if not chrome_dir.exists():
        return None

    profile_dirs = sorted(
        [path for path in chrome_dir.iterdir() if path.is_dir() and (path.name == "Default" or path.name.startswith("Profile "))]
    )
    matching_profiles = [profile for profile in profile_dirs if _profile_has_xhs_cookies(profile)]
    if not matching_profiles:
        return None
    newest_profile = max(matching_profiles, key=lambda profile: _profile_cookie_mtime(profile))
    return newest_profile.name


def _chrome_user_data_dir() -> Path:
    return Path.home() / "Library" / "Application Support" / "Google" / "Chrome"


def _profile_cookie_mtime(profile_dir: Path) -> float:
    cookie_path = _profile_cookie_path(profile_dir)
    try:
        return cookie_path.stat().st_mtime
    except OSError:
        return 0


def _profile_has_xhs_cookies(profile_dir: Path) -> bool:
    cookie_path = _profile_cookie_path(profile_dir)
    if not cookie_path.exists():
        return False

    with tempfile.TemporaryDirectory() as tmp:
        cookie_copy = Path(tmp) / "Cookies"
        try:
            shutil.copy2(cookie_path, cookie_copy)
            connection = sqlite3.connect(cookie_copy)
            try:
                row = connection.execute(
                    "select 1 from cookies where host_key like ? or host_key like ? limit 1",
                    ("%xiaohongshu.com", "%xhslink.com"),
                ).fetchone()
            finally:
                connection.close()
        except Exception:
            return False
    return row is not None


def _profile_cookie_path(profile_dir: Path) -> Path:
    network_cookie_path = profile_dir / "Network" / "Cookies"
    if network_cookie_path.exists():
        return network_cookie_path
    return profile_dir / "Cookies"


def _open_blank_window_macos() -> None:
    if _run_osascript(_chrome_blank_window_script()):
        return
    if _run_osascript(_safari_blank_window_script()):
        return
    webbrowser.open("about:blank", new=1)


def _open_url_in_current_tab_macos(url: str) -> None:
    if _run_osascript(_chrome_current_tab_script(url)):
        return
    if _run_osascript(_safari_current_tab_script(url)):
        return
    webbrowser.open(url, new=0)


def _open_url_in_new_tab_macos(url: str) -> None:
    if _run_osascript(_chrome_new_tab_script(url)):
        return
    if _run_osascript(_safari_new_tab_script(url)):
        return
    webbrowser.open_new_tab(url)


def _open_blank_window_chromium_macos(browser_name: str) -> None:
    if not _run_osascript(_chromium_blank_window_script(browser_name)):
        webbrowser.open("about:blank", new=1)


def _open_blank_window_chrome_profile_macos(profile_name: str) -> None:
    _open_url_in_new_window_chrome_profile_macos(profile_name, "about:blank")


def chrome_profile_new_window_command(profile_name: str, url: str) -> list[str]:
    return [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        f"--profile-directory={profile_name}",
        "--new-window",
        url,
    ]


def _open_url_in_new_window_chrome_profile_macos(profile_name: str, url: str) -> None:
    command = chrome_profile_new_window_command(profile_name, url)
    try:
        subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, start_new_session=True)
    except Exception:
        _open_url_in_new_window_chromium_macos("Google Chrome", url)
    time.sleep(0.7)


def _open_url_in_new_window_chromium_macos(browser_name: str, url: str) -> None:
    script = f"""
tell application {_applescript_string(browser_name)}
    activate
    make new window
    set URL of active tab of front window to {_applescript_string(url)}
end tell
""".strip()
    if not _run_osascript(script):
        webbrowser.open(url, new=1)


def _open_url_in_new_window_safari_macos(url: str) -> None:
    script = f"""
tell application "Safari"
    activate
    make new document with properties {{URL:{_applescript_string(url)}}}
end tell
""".strip()
    if not _run_osascript(script):
        webbrowser.open(url, new=1)


def _open_url_in_current_tab_chromium_macos(browser_name: str, url: str) -> None:
    if not _run_osascript(_chromium_current_tab_script(browser_name, url)):
        webbrowser.open(url, new=0)


def _open_url_in_new_tab_chromium_macos(browser_name: str, url: str) -> None:
    if not _run_osascript(_chromium_new_tab_script(browser_name, url)):
        webbrowser.open_new_tab(url)


def _open_blank_window_safari_macos() -> None:
    if not _run_osascript(_safari_blank_window_script()):
        webbrowser.open("about:blank", new=1)


def _open_url_in_current_tab_safari_macos(url: str) -> None:
    if not _run_osascript(_safari_current_tab_script(url)):
        webbrowser.open(url, new=0)


def _open_url_in_new_tab_safari_macos(url: str) -> None:
    if not _run_osascript(_safari_new_tab_script(url)):
        webbrowser.open_new_tab(url)


def _run_osascript(script: str) -> bool:
    try:
        subprocess.run(["osascript", "-e", script], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        return False
    return True


def _chromium_blank_window_script(browser_name: str) -> str:
    return f"""
tell application {_applescript_string(browser_name)}
    activate
    set matchedWindow to missing value
    repeat with browserWindow in windows
        repeat with browserTab in tabs of browserWindow
            set tabUrl to URL of browserTab
            if tabUrl contains "pgy.xiaohongshu.com" then
                set matchedWindow to browserWindow
                exit repeat
            end if
        end repeat
        if matchedWindow is not missing value then exit repeat
    end repeat
    if matchedWindow is not missing value then
        set index of matchedWindow to 1
    else if (count of windows) > 0 then
        set index of front window to 1
    end if
end tell
delay 0.2
tell application "System Events"
    keystroke "n" using command down
end tell
delay 0.5
""".strip()


def _chromium_current_tab_script(browser_name: str, url: str) -> str:
    return f"""
tell application {_applescript_string(browser_name)}
    activate
    set URL of active tab of front window to {_applescript_string(url)}
end tell
""".strip()


def _chromium_new_tab_script(browser_name: str, url: str) -> str:
    return f"""
tell application {_applescript_string(browser_name)}
    activate
    tell front window to make new tab with properties {{URL:{_applescript_string(url)}}}
end tell
""".strip()


def _chrome_blank_window_script() -> str:
    return """
tell application "Google Chrome"
    activate
    make new window
    set URL of active tab of front window to "about:blank"
end tell
""".strip()


def _chrome_current_tab_script(url: str) -> str:
    return f"""
tell application "Google Chrome"
    activate
    set URL of active tab of front window to {_applescript_string(url)}
end tell
""".strip()


def _chrome_new_tab_script(url: str) -> str:
    return f"""
tell application "Google Chrome"
    activate
    tell front window to make new tab with properties {{URL:{_applescript_string(url)}}}
end tell
""".strip()


def _safari_blank_window_script() -> str:
    return """
tell application "Safari"
    activate
    make new document with properties {URL:"about:blank"}
end tell
""".strip()


def _safari_current_tab_script(url: str) -> str:
    return f"""
tell application "Safari"
    activate
    set URL of current tab of front window to {_applescript_string(url)}
end tell
""".strip()


def _safari_new_tab_script(url: str) -> str:
    return f"""
tell application "Safari"
    activate
    tell front window to set current tab to (make new tab with properties {{URL:{_applescript_string(url)}}})
end tell
""".strip()


def _applescript_string(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def _progress(callback: Callable[[str], None] | None, message: str) -> None:
    if callback:
        callback(message)
    else:
        print(message)


def _cell_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量打开小红书蒲公英博主详情页")
    parser.add_argument("excel", help="包含小红书昵称和主页链接的 Excel 文件路径")
    parser.add_argument("--sheet", help="工作表名称；默认使用第一个工作表")
    parser.add_argument("--start", type=int, default=1, help="从第几个有效博主开始，按 1 起算；默认 1")
    parser.add_argument("--count", type=int, default=10, help="每次打开几个标签页；默认 10")
    parser.add_argument("--delay", type=float, default=0.3, help="打开标签页间隔秒数；默认 0.3")
    parser.add_argument("--window-size", type=int, default=10, help="每几个链接新开一个浏览器窗口；默认 10")
    parser.add_argument("--dry-run", action="store_true", help="只打印链接，不打开浏览器")
    parser.add_argument("--no-resolve", action="store_true", help="不解析 xhslink.com 短链")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    print(f"蒲公英博主打开工具 v{TOOL_VERSION}", flush=True)
    print(f"开始处理 Excel：从第 {args.start} 个有效博主开始，最多打开 {args.count} 个。", flush=True)
    creators = load_creators(
        args.excel,
        sheet_name=args.sheet,
        resolve_short_links=not args.no_resolve,
        start=args.start,
        count=args.count,
        progress=lambda message: print(message, flush=True),
    )
    batch = creators
    open_creator_tabs(batch, delay=args.delay, dry_run=args.dry_run, window_size=args.window_size)

    failed = [creator for creator in batch if creator.error]
    if failed:
        print("\n以下博主未能打开：")
        for creator in failed:
            print(f"- 第 {creator.row_number} 行 {creator.name}: {creator.error}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
